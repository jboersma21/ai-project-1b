"""
Project Part 1b

Team 6

Colin Moody, Ohad Beck, Charlie MacVicar, Jake Boersma

"""
import string
from openpyxl import load_workbook

DATA_FILE = 'data/Initial-World.xlsx'


def create_resource_dict():
    """
    Creates a dictionary containing resource and weight info.
    Keys: name of resource
    Values: weight of corresponding resource
    """
    wb = load_workbook(DATA_FILE)
    resources_sheet = wb.get_sheet_by_name('Resources')
    resource_dict = {}
    for row in range(2, resources_sheet.max_row + 1):
        key = get_val(resources_sheet, 'A', row)
        value = get_val(resources_sheet, 'B', row)
        resource_dict[key] = value
    return resource_dict


def create_country_dict():
    """
    Creates a dictionary containing country info.
    Keys: name of country
    Values: dictionary where key = resource_name and value = resource_amount
    """
    wb = load_workbook(DATA_FILE)
    country_sheet = wb.get_sheet_by_name('Countries')
    country_dict = {}
    for row in range(2, country_sheet.max_row + 1):
        key = get_val(country_sheet, 'A', row)
        resrc_dict = {}
        for col in range(2, country_sheet.max_column + 1):
            resrc_key = get_val(country_sheet, col_letter(col), 1)
            resrc_value = get_val(country_sheet, col_letter(col), row)
            resrc_dict[resrc_key] = resrc_value
        resrc_dict["R21'"] = 0
        resrc_dict["R22'"] = 0
        resrc_dict["R23'"] = 0
        country_dict[key] = resrc_dict
    return country_dict


def col_letter(col_num):
    letter_dict = dict(enumerate(string.ascii_uppercase, 1))
    return letter_dict[col_num]


def get_val(sheet, col, row):
    """Returns the value of a cell."""
    return sheet[col + str(row)].value


def print_resource_dict(dict):
    """Prints a resource dictionary's key and values line by line."""
    for key in dict:
        print(key, dict[key])


def print_country_dict(dict):
    """Prints a country dictionary's key and values line by line."""
    for key in dict:
        dict2 = dict[key]
        print(key)
        for key2 in dict2:
            print('\t' + key2, dict2[key2])
