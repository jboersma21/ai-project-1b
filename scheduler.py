import sys
import heapq
from openpyxl import load_workbook

import data_import
from world_objects import World
from config import configuration


# Implement state manager to traverse through of current state, future states, and previous states
class WorldStateManager(object):

    def __init__(self, depth_bound, initial_resources, initial_countries):
        self.cur_state = World(d_bound=depth_bound, weight_dict=initial_resources, country_dict=initial_countries)
        self.future_states = list()    # priority queue that store states based on big-u
        self.prev_states = list()      # stack of explored states (i.e. trajectory so far)

    # unfinished depth-bound search algorithm
    def execute_search(self):
        while self.cur_state not in self.prev_states:
            self.go_to_next_state()
            # to-do: add depth-bounded logic

    def go_to_next_state(self):
        for world in generate_successors(self.cur_state):
            self.add_future_state(world_state=world)
        # self.add_future_state(self.cur_state)
        self.prev_states.append(self.cur_state)
        self.cur_state = self.pop_future_state()

    def add_future_state(self, world_state):
        heapq.heappush(self.future_states, (world_state.get_big_u(), world_state))
        self.future_states.sort(reverse=True)

    def pop_future_state(self):
        return heapq.heappop(self.future_states)[1]

    def print_state_info(self):
        print('Current State: {}\t{}'.format(self.cur_state, self.cur_state.get_big_u()))
        print('Future States: {}'.format(self.future_states))
        print('Prev States: {}'.format(self.prev_states))


def generate_successors(current_state):
    successors = list()

    # Add every transformation for every country
    for country in current_state.countries.keys():
        for operator in configuration["transformations"]:
            tmp_world = current_state.get_deep_copy()
            if tmp_world.countries[country].transform(transformation=operator):
                successors.append(tmp_world)

    # Add every transfer for every pair of countries (both ways)
    for exporter in current_state.countries.keys():
        for destination in current_state.countries.keys():
            if exporter != destination:
                for resource in configuration["resources"]:
                    tmp_world = current_state.get_deep_copy()
                    if tmp_world.transfer(exporter=exporter, destination=destination, resource=resource):
                        successors.append(tmp_world)

    return successors


def output_successors_to_excel(file_name, successors):
    wb = load_workbook(file_name)
    sheet_name = 'Successors (Test Results)'
    if sheet_name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(sheet)
    ws = wb.create_sheet(sheet_name)
    cur_row = 1
    cur_col = 1

    for idx, state in enumerate(successors):
        ws.cell(row=cur_row, column=cur_col).value = 'Successor'
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = idx
        cur_col += 2
        ws.cell(row=cur_row, column=cur_col).value = 'Big-U'
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = state.get_big_u()

        cur_row += 1
        cur_col = 1
        ws.cell(row=cur_row, column=cur_col).value = 'Country'
        cur_col += 1
        for r in configuration['resources']:
            ws.cell(row=cur_row, column=cur_col).value = r
            cur_col += 1

        cur_row += 1
        cur_col = 1
        for name, country in state.countries.items():
            ws.cell(row=cur_row, column=cur_col).value = name
            cur_col += 1
            for r in configuration['resources']:
                ws.cell(row=cur_row, column=cur_col).value = country.resources[r]
                cur_col += 1
            cur_col = 1
            cur_row += 1

        cur_row += 1

    print('{} Test Complete'.format(file_name))
    wb.save(file_name)


def run_successor_test(file_name):
    my_state_manager = WorldStateManager(depth_bound=3,
                                         initial_resources=data_import.create_resource_dict(file_name=file_name),
                                         initial_countries=data_import.create_country_dict(file_name=file_name))
    output_successors_to_excel(file_name=file_name, successors=generate_successors(my_state_manager.cur_state))


def main(argv):
    for f_name in ['data/Initial-World-Test1.xlsx']:
        run_successor_test(f_name)


if __name__ == "__main__":
    main(sys.argv)
